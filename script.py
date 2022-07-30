from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
import os
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from bs4 import BeautifulSoup
from selenium.webdriver.common.action_chains import ActionChains

##reading file
file=openpyxl.load_workbook('data.xlsx')
sheet=file.active

##driver
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = os.environ.get("GOOGLE_CHROME_BIN")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--no-sandbox")
driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=chrome_options)

#### attempting to open New section
driver.get('https://www.google.com')
input=driver.find_element(by=By.NAME,value="q")
input.send_keys('Bot started !!!')
input.submit()
time.sleep(2)
driver.find_element(by=By.XPATH,value="//*[contains(text(), 'News')]").click()

print(' Do not close the program !!!')
print('Processing .....')

### checking for URLS Verified or not
for i in range(2,sheet.max_row+1):
        input=driver.find_element(by=By.NAME,value="q")
        input.clear()
        input.send_keys(f'site:{sheet.cell(column=1,row=i).value}')
        input.submit()
        if len(driver.find_elements(by=By.CLASS_NAME,value='iRPxbe'))==0:
                sheet.cell(column=2,row=i).value='NO'
        else:
                sheet.cell(column=2,row=i).value='YES'
        print(f'{i-1}/{sheet.max_row} URLS DONE.')
        file.save('data.xlsx')
